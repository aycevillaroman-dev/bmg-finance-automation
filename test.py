from openpyxl import load_workbook
import pandas as pd
# Load the workbook and select the active worksheet
wb = load_workbook('BMG_Outsourcing_Inc_-_Journal_Report_-_Looseleaf April 2025 - Copy.xlsx')
ws = wb.active
 
# 1. Get the values from the first header row
# Iterate over cells in a specific row and extract their values
header_row_1_values = [cell.value for cell in ws[5]]


print("Header Row 1 Values:", header_row_1_values)
